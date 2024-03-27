#Flask imports
from flask import render_template, flash, redirect, url_for, request, send_file
from flask_login import current_user, login_user, logout_user, login_required
from werkzeug.urls import url_parse
from datetime import datetime, timedelta
from sqlalchemy import and_, or_, extract
import sqlalchemy as sa
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from openpyxl.comments import Comment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from flask_mail import Mail, Message
from io import BytesIO
from app.forms import ResetPasswordRequestForm, ResetPasswordForm
from app.email import send_password_reset_email, send_verification_email
from app.decorator import requires_verification

import calendar as cal
import pandas as pd
import io
import openpyxl
import matplotlib.pyplot as plt
import base64

# Local imports
from app import app, db
from app.forms import LoginForm, RegistrationForm
from app.models import User, Car, Booking, Contacts, Groups, Renewal

mail = Mail()

@app.route('/')
def home():
    return render_template('home.html', title='Home')

                                                                                # Users Login/Logout/Profile/Admin

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated: 
        return redirect(url_for('overview')) 
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password', "error")
            return redirect(url_for('login'))
        login_user(user, remember=form.remember_me.data)
        next_page = request.args.get('next')
        if not next_page or url_parse(next_page).netloc != '':
            next_page = url_for('overview')
        return redirect(next_page)
    return render_template('login.html', title='Sign In', form=form)

@app.route('/register', methods=['GET', 'POST']) 
def register():
    if current_user.is_authenticated: 
        return redirect(url_for('overview'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()

        # Send verification email
        send_verification_email(user)

        flash('Please check and confirm your email to continue!')
        
        # Log in the new user
        login_user(user)
        
        return redirect(url_for('confirm_email'))
    return render_template('register.html', title='Register', form=form)

@app.route('/confirm_email', methods=['GET', 'POST']) 
def confirm_email():
    if current_user.is_verified:
        return redirect(url_for('overview'))
    else:
        return render_template('confirm_email.html', title='Confirm Email', current_user=current_user,
                               user_name=current_user.username if current_user.is_authenticated else None)
    
@app.route('/send_verification_email',  methods=['GET', 'POST']) 
def send_verification_email_route():
    if not current_user.is_verified:
        send_verification_email(current_user)
        return render_template('confirm_email.html', title='Confirm Email', current_user=current_user,
                                user_name=current_user.username if current_user.is_authenticated else None)  
    else:
        return redirect(url_for('overview'))
  
@app.route('/confirm_email/<token>', methods=['GET'])
def confirm_email_token(token):
    user = User.verify_verification_token(token)
    print('User.verify_verification_token(token)', User.verify_verification_token(token))
    if user:
        # Mark the user as verified
        user.is_verified = True
        db.session.commit()
        flash('Your email address has been successfully verified!', 'success')
        return redirect(url_for('profile'))  
    else:
        flash('Invalid or expired token. Please try again.', 'error')
        return redirect(url_for('confirm_email'))

@app.route('/reset_password_request', methods=['GET', 'POST'])
def reset_password_request():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = ResetPasswordRequestForm()
    if form.validate_on_submit():
        user = db.session.scalar(
            sa.select(User).where(User.email == form.email.data))
        if user:
            send_password_reset_email(user)
        flash('Check your email for the instructions to reset your password')
        return redirect(url_for('login'))
    return render_template('reset_password_request.html',
                           title='Reset Password', form=form)

@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    user = User.verify_reset_password_token(token)
    if not user:
        return redirect(url_for('index'))
    form = ResetPasswordForm()
    if form.validate_on_submit():
        user.set_password(form.password.data)
        db.session.commit()
        flash('Your password has been reset.')
        return redirect(url_for('login'))
    return render_template('reset_password.html', form=form)

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('overview'))

@app.route('/profile', methods=['GET', 'POST'])
@requires_verification
def profile(): 
    current_datetime = datetime.utcnow()
    time_difference = current_datetime - current_user.registration_date
    one_minute = timedelta(minutes=1)
    if request.method == 'POST':
        if 'username' in request.form:
            new_username = request.form.get('username')
            existing_username = User.query.filter_by(username=new_username).first()
            if existing_username:
                flash(f'{existing_username.username} already taken, please choose another one', 'error')
                return redirect(url_for('profile'))
            else:
                current_user.change_username(new_username)
                db.session.commit()
                flash('Username successfully changed!', 'success')
                return redirect(url_for('profile'))                
                
        elif 'email' in request.form:
            new_email = request.form.get('email')
            existing_email = User.query.filter_by(email=new_email).first()
            if existing_email:
                flash(f'{existing_email.email} already taken, please choose another one', 'error')
                return redirect(url_for('profile'))
            else:
                current_user.email = new_email
                db.session.commit()
                flash('Email successfully changed!', 'success')
                return redirect(url_for('profile'))   
            
        elif 'password' in request.form:        
            new_password = request.form.get('password')
            password_confirm = request.form.get('confirm_password')

            if new_password == password_confirm:
                current_user.set_password(new_password)
                db.session.commit()
                flash('Password successfully changed!', 'success')
                return redirect(url_for('profile'))  
            else:
                flash('Passwords do not match!', 'error') 
                return redirect(url_for('profile'))  
            
        elif 'delete_account' in request.form:
            user_to_delete = User.query.get(current_user.id)
            if user_to_delete:
                user_to_delete.delete_user()
                flash('User succefully deleted!', 'success')
                return redirect(url_for('login'))
            else:
                return flash('User not found!', 'error')  

        elif 'currency' in request.form:
            currency = request.form.get('currency')
            language = request.form.get('language')
            unit = request.form.get('unit')

            current_user.change_preferences(currency, unit, language)
            flash('Preferences updated!', 'success')

    return render_template('profile.html' if current_user.language == 'en' else f'profile_{current_user.language}.html', 
                           time_difference=time_difference, one_minute=one_minute, title='Profile', page='profile', 
                           user=current_user, current_datetime=current_datetime, 
                           user_name=current_user.username if current_user.is_authenticated else None)


@app.route('/admin', methods=['GET', 'POST'])
@requires_verification
def admin(): 
    if current_user.username == 'admin' and current_user.role == 'admin':
        users = User.query.all()
        cars = Car.query.all()
        bookings = Booking.query.all()
        revenue = 0
        for booking in bookings:
            revenue += booking.money
        return render_template('admin.html', page='admin', total_users=len(users), total_cars=len(cars), total_bookings=len(bookings), total_revenue=revenue)
    else:
        return render_template('404.html')
    
@app.route('/users_list', methods=['GET', 'POST'])
@requires_verification
def users_list(): 
    if request.method == 'POST':
        user_id = request.form.get('activation') #Change the is_verified user variable
        user = User.query.filter_by(id=user_id).one()
        if user:
            if user.is_verified == True:
                user.is_verified = False
                db.session.commit()
                flash(f'The user with ID {user.id} and Email {user.email} has been Deactivated', 'success')
                return redirect(url_for('users_list'))
            else:
                user.is_verified = True
                db.session.commit()
                flash(f'The user with ID {user.id} and Email {user.email} has been Activated', 'success')
                return redirect(url_for('users_list'))
    if current_user.username == 'admin' and current_user.role == 'admin':
        users = User.query.all()
        return render_template('users_list.html', page='users_list', users=users)
    else:
        return render_template('404.html')
    
@app.route('/user/<username>', methods=['GET', 'POST'])
@requires_verification
def user(username):
    if current_user.username == 'admin' and current_user.role == 'admin':
        user = User.query.filter_by(username=username).one()
        cars = Car.query.filter_by(user_id=user.id)
        bookings = Booking.query.filter_by(user_id=user.id)
        if request.method == 'POST':
            user_id = request.form.get('user_id')
            user_to_delete = User.query.get(user_id)
            if user_to_delete:
                user_to_delete.delete_user()
                flash('User succefully deleted!', 'success')
                return redirect(url_for('users_list'))
            else:
                return flash('User not found!', 'error')          
        return render_template('user.html', page='users_page', user=user, cars=cars, bookings=bookings)        
    else:
        return render_template('404.html')

                                                                                # Garage

@app.route('/garage_view', methods=['GET', 'POST'])
@requires_verification
def garage_view():
    user_cars = current_user.garage.all()
    return render_template('garage_view.html' if current_user.language == 'en' else f'garage_view_{current_user.language}.html', 
                           title='Garage', 
                           page="garage_view", user_cars=user_cars, user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/garage_manage', methods=['GET', 'POST'])
@requires_verification
def garage_manage(): # Add Car
    user_cars = current_user.garage.all()
    if request.method == 'POST':
       # Process the form data for a POST request
       plate = request.form.get('Plate').upper()
       make = request.form.get('Make').lower().capitalize()
       model = request.form.get('Model').lower().capitalize()
       fuel = request.form.get('Fuel').lower().capitalize()
       year = request.form.get('Year').lower().capitalize()
       cc = request.form.get('Cc').lower().capitalize()

       # Check if a car with the same plate and user_id already exists
       existing_car = Car.query.filter_by(plate=plate).first()

       if existing_car is not None:
           flash('A car with this plate already exists in the database', 'error')
           return render_template('garage_manage.html' if current_user.language == 'en' else f'garage_manage_{current_user.language}.html',
                              user_name=current_user.username if current_user.is_authenticated else None)

       new_car = Car(plate=plate, make=make, model=model, fuel=fuel, year=year, cc=cc, user_id=current_user.id)

       db.session.add(new_car)
       db.session.commit()
       flash('Car added successfully', 'success')
       return render_template('garage_manage.html' if current_user.language == 'en' else f'garage_manage_{current_user.language}.html',
                              user_name=current_user.username if current_user.is_authenticated else None)
    else:
       # Render the form for a GET request
       return render_template('garage_manage.html' if current_user.language == 'en' else f'garage_manage_{current_user.language}.html', 
                              title='Add Car', page="garage_manage", user_cars=user_cars,
                              user_name=current_user.username if current_user.is_authenticated else None)


@app.route('/search', methods=['GET', 'POST'])
def search():
    user_cars = current_user.garage.all()
    current_page = request.form.get('source_page', default='garage_manage' if current_user.language == 'en' else f'garage_manage_{current_user.language}')
    

    if request.method == 'POST':
        search_query = request.form.get('search_query')
        search_type = request.form.get('search_type')
        select_car = request.form.get('select_car') # Plate

        # Perform search based on both search type and selected car
        if search_type and search_query:
            filtered_cars = Car.search(search_query, search_type, current_user.id)
        elif select_car and select_car != "blank":
            # If a car is selected, perform search based on the selected car
            filtered_cars = [Car.query.filter_by(plate=select_car).first()]
        else:
            # No valid search parameters provided
            filtered_cars = []

        if current_page == 'garage_manage' or current_page == f'garage_manage_{current_user.language}':
            return render_template('garage_manage.html' if current_user.language == 'en' else f'garage_manage_{current_user.language}.html', 
                                   cars=filtered_cars, search_type=search_type, search_query=search_query, user_cars=user_cars,
                                   user_name=current_user.username if current_user.is_authenticated else None)
        elif current_page == 'garage_car':
            if select_car != "blank":
                current_datetime = datetime.utcnow()
                # Retrieve active bookings for a given car
                car_active_bookings = Booking.query.filter(
                    (Booking.car_plate == select_car) &
                    (Booking.end_datetime > current_datetime)
                ).all()

                # Retrieve past bookings for a given car
                car_past_bookings = Booking.query.filter(
                    (Booking.car_plate == select_car) &
                    (Booking.end_datetime <= current_datetime)
                ).all()

                car = Car.query.filter_by(plate=select_car).first()
                return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                        title='Car', page="garage_car", user_cars=user_cars,
                                        car_object=car, active_bookings=car_active_bookings, past_bookings=car_past_bookings,
                                        user_name=current_user.username if current_user.is_authenticated else None)
            return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                   cars=filtered_cars, search_type=search_type, search_query=search_query, user_cars=user_cars,
                                   user_name=current_user.username if current_user.is_authenticated else None)
        elif current_page == 'downloads':
            return render_template('downloads.html' if current_user.language == 'en' else f'downloads_{current_user.language}.html', 
                                   cars=filtered_cars, search_type=search_type, search_query=search_query, user_cars=user_cars,
                                   single_data_button='car',
                                   user_name=current_user.username if current_user.is_authenticated else None)
    
    else:
        # Render the search page template for a GET request
        return render_template('garage_manage.html' if current_user.language == 'en' else f'garage_manage_{current_user.language}.html',
                               user_name=current_user.username if current_user.is_authenticated else None)
    
@app.route('/delete', methods=['POST'])
def delete_car():
    car_plate = request.form.get('car_plate')
    if car_plate:
        car = Car.query.filter_by(plate=car_plate).first()
        if car:
            
            # Remove bookings associated with the car
            bookings_to_remove = Booking.query.filter_by(car_plate=car_plate).all()
            for booking in bookings_to_remove:
                Booking.remove_booking(booking.id)

            # Delete the car
            Car.delete_car(car_plate)
            flash('Car deleted successfully.', 'success')
        else:
            flash('Car not found.', 'error')
    else:
        flash('Invalid request.', 'error')
    return redirect(url_for('garage_view'))

@app.route('/overview', methods=['GET', 'POST'])
@requires_verification
def overview(): # Booking
    contact_id = request.args.get('contact_id')
    contact = None
    if contact_id:
        contact = Contacts.query.filter_by(id=contact_id).first() 

    if request.method == 'POST':
        try:
            if 'book' in request.form:
                car_plate = request.form.get('car_selection')
                price = int(request.form.get('Price'))
                group_id = request.form.get('group_selection')
                km = int(request.form.get('km'))
                start_date = request.form.get('start_date')
                end_date = request.form.get('end_date')
                start_time = request.form.get('start_time')
                end_time = request.form.get('end_time')
                note = request.form.get('note')
                full_name = request.form.get("full_name")
                dob = request.form.get("dob")
                driver_licence_n = request.form.get("driver_licence_n")                
                telephone = request.form.get("telephone") 
                contact_id = request.form.get("id")
                
                # Convert start, end, to and from date and time to a datetime object
                start_datetime = datetime.strptime(f'{start_date} {start_time}', '%Y-%m-%d %H:%M')
                end_datetime = datetime.strptime(f'{end_date} {end_time}', '%Y-%m-%d %H:%M')

                # Check if start_datetime is in the past
                if start_datetime < datetime.now():
                    flash('Booking cannot be made for a past date and time.', 'error')
                    return redirect(url_for('overview'))

                if contact_id == None:
                    if full_name == '': # Skip if client's fields are empty
                        pass
                    else:
                        # Create a new contact
                        new_contact = Contacts(
                        full_name=full_name,
                        driver_licence_n=driver_licence_n,
                        dob=dob,
                        telephone=telephone,
                        user_id=current_user.id
                        )
                        db.session.add(new_contact)
                        db.session.commit()

                        #Get the id of the newly created contact
                        contact_id = new_contact.id                

                # Call the create_booking method from the Booking model
                booking, overlap_start, overlap_end = Booking.create_booking(
                    car_plate, price, start_datetime, end_datetime, contact_id, current_user.id, note, km, group_id
                )
                
                if booking is None:
                    flash(f'This car is already booked from {overlap_start.strftime("%b %d %H:%M")} to {overlap_end.strftime("%b %d %H:%M")}!', 'error')
                else:
                    flash('Car booked successfully!', 'success')
                return redirect(url_for('overview'))  # Redirect after a successful form submission

            elif 'check' in request.form:
                # To check available and booked cars
                from_date = request.form.get('from')
                to_date = request.form.get('to')                

        except ValueError as e:
            flash(str(e), 'error')
            print(f'ValueError: {str(e)}')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
            print(f'Error: {str(e)}')

    user_cars = current_user.garage.all()
    user_contacts = current_user.contacts.all()
    user_groups = current_user.groups.all()

    # To check available and booked cars
    from_date = request.form.get('from')
    to_date = request.form.get('to')

    # Standard time frame if no Input
    if from_date == None:
        from_date = datetime.now().strftime('%Y-%m-%d')
        to_date = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
        from_datetime = datetime.strptime(f'{from_date}', '%Y-%m-%d')
        to_datetime = datetime.strptime(f'{to_date}', '%Y-%m-%d')
    # Chosen time frame if Input    
    else:
        from_datetime = datetime.strptime(f'{from_date}', '%Y-%m-%d')
        to_datetime = datetime.strptime(f'{to_date}', '%Y-%m-%d')
        #Include the full to_datetime
        to_datetime += timedelta(days=1)    

    # Fetch all bookings for the user within the specified time range
    user_bookings = Booking.query.filter(
                Booking.user_id == current_user.id,
                Booking.end_datetime >= from_datetime,
                Booking.start_datetime <= to_datetime,
                ).all()

    # Extract the plates of booked cars
    booked_car_plates = [booking.car_plate for booking in user_bookings]

    # Filter out booked cars from available cars
    available_cars = [car for car in user_cars if car.plate not in booked_car_plates]

    # Fetch bookings about to finish    
    bookings_about_to_finish = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        ((Booking.end_datetime >= datetime.now().date()) & (Booking.end_datetime <= (datetime.now().date() + timedelta(weeks=1))))).all()
    
    # Fetch bookings about to start    
    bookings_about_to_start = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        ((Booking.start_datetime >= datetime.now().date()) & (Booking.start_datetime <= (datetime.now().date() + timedelta(weeks=1))))).all()

    # Fetch cars with expiring insurance, MOT, and road tax in a month
    in_four_weeks = datetime.now() + timedelta(weeks=4)
    cars_expiring_soon = Car.query.filter(
        (Car.user_id == current_user.id) &
        ((Car.insurance_expiry_date <= in_four_weeks) |
        (Car.mot_expiry_date <= in_four_weeks) |
        (Car.road_tax_expiry_date <= in_four_weeks))
    ).all()

    # Render the template with the new data
    return render_template('overview.html' if current_user.language == 'en' else f'overview_{current_user.language}.html',
                            user_cars=user_cars, user_groups=user_groups,
                            available_cars=available_cars,
                            user_bookings=user_bookings,
                            from_datetime=from_datetime.strftime('%Y-%m-%d'),
                            to_datetime=to_datetime.strftime('%Y-%m-%d'),
                            user_contacts=user_contacts,
                            contact=contact,
                            user_name=current_user.username if current_user.is_authenticated else None,
                            bookings_about_to_start=bookings_about_to_start,
                            bookings_about_to_finish=bookings_about_to_finish,
                            in_four_weeks=in_four_weeks,
                            cars_expiring_soon=cars_expiring_soon)
    
@app.route('/renew', methods=['GET', 'POST'])
@requires_verification
def renew():
    car_plate = request.args.get('car_plate')
    car_object = Car.query.filter_by(plate=car_plate).first()
    current_datetime = datetime.utcnow()

    if 'select_cost' in request.form:
        selected_option = request.form['select_cost']
        return render_template('renew.html' if current_user.language == 'en' else f'renew_{current_user.language}.html', 
                               renewals=car_object.renewal.all(), option=selected_option, car_object=car_object, 
                               user_name=current_user.username if current_user.is_authenticated else None)
    
    elif 'cost_amount' in request.form:
        try:
            cost_amount = int(request.form['cost_amount'])
        except ValueError:
            # Handle the case where the input is not an integer
            flash('Cost amount must be an integer.', 'error')
            # Redirect or render the template with an error message
            return redirect(url_for('renew', car_plate=car_object.plate))
        option = request.form['option']

        # Fetch the forms
        insurance_expiry_date_str = request.form.get('insurance_expiry_date')
        road_tax_expiry_date_str = request.form.get('road_tax_expiry_date')
        mot_expiry_date_str = request.form.get('mot_expiry_date')
        other = request.form.get('other')
        note = request.form['note']

        # Convert date strings to datetime objects
        road_tax_expiry_date = datetime.strptime(road_tax_expiry_date_str, '%Y-%m-%d') if road_tax_expiry_date_str else None
        mot_expiry_date = datetime.strptime(mot_expiry_date_str, '%Y-%m-%d') if mot_expiry_date_str else None
        insurance_expiry_date = datetime.strptime(insurance_expiry_date_str, '%Y-%m-%d') if insurance_expiry_date_str else None
        
        # Create a variable to pass to car_object.add_renewal
        if road_tax_expiry_date_str:
            new_expiry_date = road_tax_expiry_date
        elif mot_expiry_date:
            new_expiry_date = mot_expiry_date
        elif insurance_expiry_date:
            new_expiry_date = insurance_expiry_date
        elif other:
            new_expiry_date = current_datetime
    
        car_object.add_renewal(renewal_type=option, renewal_date=new_expiry_date, renewal_cost=cost_amount, current_datetime=current_datetime, description=note)
        flash(f'{option.capitalize()} expiry date has been updated!', 'success')
        car_renewals = car_object.renewal.all()

    elif 'delete' in request.form:
        renewal_id = request.form['delete']
        car_object.delete_renewal(renewal_id)
        flash('Renewal correctly deleted!', 'success')
        car_renewals = car_object.renewal.all()

    else:
        car_renewals = car_object.renewal.all()

    return render_template('renew.html' if current_user.language == 'en' else f'renew_{current_user.language}.html', 
                           renewals=car_renewals, car_object=car_object, user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/garage_car', methods=['GET', 'POST'])
@requires_verification
def garage_car():
    user_cars = current_user.garage.all()
    current_datetime = datetime.utcnow()

    # Handles the redirection from other pages
    car_plate = request.args.get('car_plate')
    if car_plate:
        # Retrieve active bookings for a given car
                car_active_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime > current_datetime)
                ).all()

                # Retrieve past bookings for a given car
                car_past_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime <= current_datetime)
                ).all()

                selected_car = Car.query.filter_by(plate=car_plate).first()
                return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                       title='Car', page="garage_car", user_cars=user_cars, car_object=selected_car,
                                        active_bookings=car_active_bookings, past_bookings=car_past_bookings,
                                        user_name=current_user.username if current_user.is_authenticated else None)

    if request.method == 'POST':
        try:
            if 'plate' in request.form:
                car_plate = request.form.get('plate')

                # Retrieve active bookings for a given car
                car_active_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime > current_datetime)
                ).all()

                # Retrieve past bookings for a given car
                car_past_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime <= current_datetime)
                ).all()

                car = Car.query.filter_by(plate=car_plate).first()
                return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                        title='Car', page="garage_car", user_cars=user_cars,
                                        car_object=car, active_bookings=car_active_bookings, past_bookings=car_past_bookings,
                                        user_name=current_user.username if current_user.is_authenticated else None)
            
            elif request.form.get('action') == 'amend':
                car_plate = request.form['car_plate']
                selected_car = Car.query.filter_by(plate=car_plate).first()

                # Retrieve active bookings for a given car
                car_active_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime > current_datetime)
                ).all()

                # Retrieve past bookings for a given car
                car_past_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime <= current_datetime)
                ).all()

                if selected_car:
                    try:
                        # Extracting form data
                        car_make = request.form['car_make']
                        car_model = request.form['car_model']
                        car_fuel = request.form['car_fuel']
                        car_year = request.form['car_year']
                        car_cc = request.form['car_cc']

                        # Call the amend_car method
                        selected_car.amend_car(car_plate, car_make, car_model, car_fuel, car_year, car_cc)

                        flash('Car amended successfully!', 'success')
                        return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                                title='Car', page="garage_car", user_cars=user_cars, car_object=selected_car,
                                                active_bookings=car_active_bookings, past_bookings=car_past_bookings,
                                                user_name=current_user.username if current_user.is_authenticated else None)

                    except ValueError as e:
                        flash(str(e), 'error')  # Handle any parsing errors
                    except Exception as e:
                        flash(f'Error: {str(e)}', 'error')  # Handle other exceptions

            elif request.form.get('action') == 'delete':
                car_plate = request.form['car_plate']
                selected_car = Car.query.filter_by(plate=car_plate).first()
                if selected_car:
                    try:
                        # Remove bookings associated with the car
                        bookings_to_remove = Booking.query.filter_by(car_plate=car_plate).all()
                        for booking in bookings_to_remove:
                            Booking.remove_booking(booking.id)

                        # Delete the car
                        Car.delete_car(car_plate)
                        flash('Car deleted successfully.', 'success')

                    except ValueError as e:
                        flash(str(e), 'error')  # Handle any parsing errors
                    except Exception as e:
                        flash(f'Error: {str(e)}', 'error')  # Handle other exceptions

            # Handles the "Manage" buttons in other pages
            elif request.form.get('garage_car'):
                car_plate = request.form.get('garage_car')

                # Retrieve active bookings for a given car
                car_active_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime > current_datetime)
                ).all()

                # Retrieve past bookings for a given car
                car_past_bookings = Booking.query.filter(
                    (Booking.car_plate == car_plate) &
                    (Booking.end_datetime <= current_datetime)
                ).all()

                selected_car = Car.query.filter_by(plate=car_plate).first()
                return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                                        title='Car', page="garage_car", user_cars=user_cars, car_object=selected_car,
                                        active_bookings=car_active_bookings, past_bookings=car_past_bookings,
                                        user_name=current_user.username if current_user.is_authenticated else None)
                
        except ValueError as e:
            flash(str(e), 'error')
            print(f'ValueError: {str(e)}')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
            print(f'Error: {str(e)}')

    return render_template('garage_car.html' if current_user.language == 'en' else f'garage_car_{current_user.language}.html', 
                           title='Car', page="garage_car", user_cars=user_cars,
                           user_name=current_user.username if current_user.is_authenticated else None)

                                                                                # Bookings 


@app.route('/bookings_view', methods=['GET', 'POST'])
@requires_verification
def bookings_view(): 
    current_date = datetime.utcnow()
    
    # Retrieve active bookings
    active_bookings = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        (Booking.end_datetime > current_date)
    ).all()

    return render_template('bookings_view.html' if current_user.language == 'en' else f'bookings_view_{current_user.language}.html',
                           user_bookings=active_bookings, page='bookings_view',
                           user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/bookings_manage', methods=['GET', 'POST'])
@requires_verification
def bookings_manage():
    current_date = datetime.utcnow()
    booking_id = request.args.get('booking_id')

    # Retrieve active bookings
    user_bookings = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        (Booking.end_datetime > current_date)
    ).all()

    if booking_id:
        selected_booking = Booking.query.filter_by(id=booking_id).first()
        contact = Contacts.query.filter_by(id=selected_booking.contact_id).first()
        group = Groups.query.filter_by(id=selected_booking.group_id).first()
        if selected_booking.is_expired():
            flash(f'Booking with ID {selected_booking.id} is no longer active, you can only delete it', 'error')
            return redirect(url_for('bookings_history'))
    else:
        selected_booking = None
        contact = None
        group = None

    if request.method == 'POST':
        booking_id = request.form.get('search_type')
        if booking_id:
            selected_booking = Booking.query.filter_by(id=booking_id).first()
            contact = Contacts.query.filter_by(id=selected_booking.contact_id).first()
            group = Groups.query.filter_by(id=selected_booking.group_id).first()
        
        if request.form.get('action') == 'delete' or 'delete' in request.form:
            booking_id = request.form.get('delete')
            # Modify the query to eagerly load the 'car' relationship
            selected_booking = Booking.query.options(db.joinedload(Booking.car)).filter_by(id=booking_id).first()
            if selected_booking:
                Booking.remove_booking(booking_id)
                flash('Booking successfully deleted!', 'success')

                # Redirect to the same page to refresh
                return redirect(url_for('bookings_view'))

        # Handles the "Manage" buttons in other pages
        if request.form.get('manage_booking'):
            booking_id = request.form.get('manage_booking')
            selected_booking = Booking.query.filter_by(id=booking_id).first()

            if selected_booking:
                contact = Contacts.query.filter_by(id=selected_booking.contact_id).first()
                group = Groups.query.filter_by(id=selected_booking.group_id).first()
            else:
                flash('Selected booking not found.', 'error')            

        # Check if the form was submitted for amendment
        if request.form.get('action') == 'amend':
            booking_id = request.form['booking_id']
            selected_booking = Booking.query.filter_by(id=booking_id).first()
            # Handle the amendment logic here, e.g., update the database
            if selected_booking:
                try:
                    # Extracting form data
                    start_date = request.form['start_date']
                    start_time = request.form['start_time']
                    end_date = request.form['end_date']
                    end_time = request.form['end_time']
                    note = request.form['note']

                    # Parsing form data to create datetime objects
                    start_datetime = datetime.strptime(f'{start_date} {start_time}', '%Y-%m-%d %H:%M')
                    end_datetime = datetime.strptime(f'{end_date} {end_time}', '%Y-%m-%d %H:%M')

                    # Call the amend_booking method
                    selected_booking.amend_booking(start_datetime, end_datetime, note)

                    flash('Booking amended successfully!', 'success')

                except ValueError as e:
                    flash(str(e), 'error')  # Handle any parsing errors
                except Exception as e:
                    flash(f'Error: {str(e)}', 'error')  # Handle other exceptions

            else:
                flash('Booking not found. Amendment failed.', 'error')
                
    return render_template('bookings_manage.html' if current_user.language == 'en' else f'bookings_manage_{current_user.language}.html',
                           page='bookings_manage', user_bookings=user_bookings, selected_booking=selected_booking, contact=contact, group=group,
                           user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/bookings_history', methods=['GET', 'POST'])
@requires_verification
def bookings_history(): 
    current_date = datetime.utcnow()
    
    # Retrieve expired bookings
    expired_bookings = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        (Booking.end_datetime < current_date)
    ).all()

    if request.method == 'POST':
        if request.form.get('delete'):
            booking_id = request.form.get('delete')
            # Modify the query to eagerly load the 'car' relationship
            selected_booking = Booking.query.options(db.joinedload(Booking.car)).filter_by(id=booking_id).first()
            if selected_booking:
                Booking.remove_booking(booking_id)
                flash('Booking successfully deleted!', 'success')

                # Redirect to the same page to refresh
                return redirect(url_for('bookings_history'))
            
        if request.form.get('delete_car'):
            booking_id = request.form.get('delete_car')
            # Modify the query to eagerly load the 'car' relationship
            selected_booking = Booking.query.options(db.joinedload(Booking.car)).filter_by(id=booking_id).first()
            if selected_booking:
                Booking.remove_booking(booking_id)
                flash('Booking successfully deleted!', 'success')

                # Redirect to the same page to refresh
                return redirect(url_for('garage_car'))
            
        if 'search' in request.form:
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            searched_booking = []
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')

            #Include the full end_date
            end_date = end_date + timedelta(days=1)
            
            # Check booking which end date and start date is within the search
            for booking in expired_bookings:
                if booking.end_datetime >= start_date and booking.end_datetime <= end_date:
                    searched_booking.append(booking)
                elif booking.start_datetime >= start_date and booking.start_datetime <= end_date:
                    searched_booking.append(booking)

            return render_template('bookings_history.html', user_bookings=expired_bookings, page='bookings_history', bookings=searched_booking,
                                   user_name=current_user.username if current_user.is_authenticated else None)

    return render_template('bookings_history.html' if current_user.language == 'en' else f'bookings_history_{current_user.language}.html',
                        user_bookings=expired_bookings, page='bookings_history',
                           user_name=current_user.username if current_user.is_authenticated else None)

                                                                                # Calendar

@app.route('/calendar', methods=['GET', 'POST'])
@requires_verification
def calendar(): 
    current_date = datetime.now()
    current_day = datetime.now().day
    current_month = datetime.now().strftime("%B %Y")
    first_day_of_month = current_date.replace(day=1)
    last_day_of_month = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    days_in_month = cal.monthrange(datetime.now().year, datetime.now().month)[1]
    is_current_month = True

    # Retrieve user's cars
    user_cars = current_user.garage.all()  
    # Query user's bookings
    user_bookings = Booking.query.filter(
        (Booking.user_id == current_user.id) &
        (
            (Booking.start_datetime <= last_day_of_month) |  # Start date is within the current month
            (Booking.end_datetime >= first_day_of_month)    # End date is within the current month
        )
    ).all()

    # Initialize a dictionary to store booking information for each day
    booking_data = defaultdict(lambda: defaultdict(list))
    
    # Populate booking data
    for booking in user_bookings:
        start_day = booking.start_datetime.day
        end_day = booking.end_datetime.day
        car_plate = booking.car_plate
        if end_day < start_day:
            end_day = current_date.replace(day=days_in_month).day
        for day in range(start_day, end_day + 1):  # Include end_day in the range
            # Check if the day is within the current month
            if first_day_of_month <= booking.start_datetime <= last_day_of_month:
                booking_data[day][car_plate].append(booking)  # Store booking objects

    if 'next_month' in request.form:
        page_date = datetime.strptime(request.form.get('next_month'), "%B %Y") # Previous month
        current_date = page_date + relativedelta(months=1) # Date of the page loaded after pressing next_month
        current_month = current_date.strftime("%B %Y")

        if current_month == datetime.now().strftime("%B %Y"): # Check if next_month is the real current month
            pass
        else:
            is_current_month = False
 
        first_day_of_month = current_date.replace(day=1)
        last_day_of_month = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = cal.monthrange(current_date.year, current_date.month)[1]
        # Retrieve user's cars
        user_cars = current_user.garage.all()  
        # Query user's bookings
        user_bookings = Booking.query.filter(
            (Booking.user_id == current_user.id) &
                ((extract('month', Booking.start_datetime) == last_day_of_month.month) |  # Start date is within the next month
                (extract('month', Booking.end_datetime) == first_day_of_month.month)    # End date is within the next month
            )
            ).all()

        # Initialize a dictionary to store booking information for each day
        booking_data = defaultdict(lambda: defaultdict(list))
        
        # Populate booking data
        for booking in user_bookings:
            start_date = booking.start_datetime.date()
            end_date = booking.end_datetime.date()
            car_plate = booking.car_plate
            
            # Check if any part of the booking falls within the next month
            if start_date <= last_day_of_month.date() and end_date >= first_day_of_month.date():
                start_day = max(start_date.day, 1)
                end_day = min(end_date.day, days_in_month)
                # If booking span in different months
                if start_day > end_day: # Ex 13 and 9
                    if start_date.month < end_date.month: 
                        if current_date.month == end_date.month:
                            start_day = 1  
                        else:
                            end_day = last_day_of_month.day
                    else: # If booking ends next month
                            end_day = last_day_of_month.day

                # If start and end are same days on different months
                elif start_day == end_day and start_date.month != end_date.month : 
                    # If the month showed is the same as the end of the booking and the booking started previous month
                    if start_date.month < end_date.month and current_date.month == end_date.month:
                        start_day = 1  
                    # if the booking ends next month
                    else:
                        end_day = last_day_of_month.day
                
                for day in range(start_day, end_day + 1):
                    booking_data[day][car_plate].append(booking)

    if 'prev_month' in request.form:
        page_date = datetime.strptime(request.form.get('prev_month'), "%B %Y")
        current_date = page_date - relativedelta(months=1)
        current_month = current_date.strftime("%B %Y")

        if current_month == datetime.now().strftime("%B %Y"): # Check if next_month is the real current month
            pass
        else:
            is_current_month = False
 
        first_day_of_month = current_date.replace(day=1)
        last_day_of_month = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = cal.monthrange(current_date.year, current_date.month)[1]
        # Retrieve user's cars
        user_cars = current_user.garage.all()  
        # Query user's bookings
        user_bookings = Booking.query.filter(
            (Booking.user_id == current_user.id) &
                ((extract('month', Booking.start_datetime) == last_day_of_month.month) |  # Start date is within the next month
                (extract('month', Booking.end_datetime) == first_day_of_month.month)    # End date is within the next month
            )
            ).all()

        # Initialize a dictionary to store booking information for each day
        booking_data = defaultdict(lambda: defaultdict(list))
        
        # Populate booking data
        for booking in user_bookings:
            start_date = booking.start_datetime.date()
            end_date = booking.end_datetime.date()
            car_plate = booking.car_plate
            
            # Check if any part of the booking falls within the next month
            if start_date <= last_day_of_month.date() and end_date >= first_day_of_month.date():
                start_day = max(start_date.day, 1)
                end_day = min(end_date.day, days_in_month)
                # If booking span in different months
                if start_day > end_day: # Ex 13 and 9
                    if start_date.month < end_date.month: 
                        if current_date.month == end_date.month:
                            start_day = 1  
                        else:
                            end_day = last_day_of_month.day
                    else: # If booking ends next month
                            end_day = last_day_of_month.day

                # If start and end are same days on different months
                elif start_day == end_day and start_date.month != end_date.month : 
                    # If the month showed is the same as the end of the booking and the booking started previous month
                    if start_date.month < end_date.month and current_date.month == end_date.month:
                        start_day = 1  
                    # if the booking ends next month
                    else:
                        end_day = last_day_of_month.day
                
                for day in range(start_day, end_day + 1):
                    booking_data[day][car_plate].append(booking)

    if 'calendar_download' in request.form:
        current_date = datetime.strptime(request.form.get('current_date'), "%Y-%m-%d %H:%M:%S")
        first_day_of_month = current_date.replace(day=1)
        last_day_of_month = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        days_in_month = cal.monthrange(current_date.year, current_date.month)[1]
        cars = Car.query.filter_by(user_id=current_user.id).all()

        # Retrieve user's cars
        user_cars = current_user.garage.all()
        # Query user's bookings
        user_bookings = Booking.query.filter(
            (Booking.user_id == current_user.id) &
            (
                (Booking.start_datetime <= last_day_of_month) |  # Start date is within the current month
                (Booking.end_datetime >= first_day_of_month)    # End date is within the current month
            )
        ).all()

        # Initialize a dictionary to store booking information for each day
        booking_data = defaultdict(lambda: defaultdict(list))

        # Populate booking data
        for booking in user_bookings:
            start_day = booking.start_datetime.day
            end_day = booking.end_datetime.day
            car_plate = booking.car_plate
            if end_day < start_day:
                end_day = current_date.replace(day=days_in_month).day
            for day in range(start_day, end_day + 1):  # Include end_day in the range
                # Check if the day is within the current month
                if first_day_of_month <= booking.start_datetime <= last_day_of_month:
                    booking_data[day][car_plate].append(booking)  # Store booking objects

        # Initialize lists to store data
        car_data = []

        for car in cars:
            car_row = [f"{car.model} - {car.plate}"]
            for day in range(1, int(days_in_month) + 1):
                if booking_data[day][car.plate]:
                    booking = booking_data[day][car.plate]  
                    if len(booking) == 1:
                        car_row.append(booking[0].id)
                    if len(booking) > 1:
                        car_row.append('M')
                else:
                    car_row.append('')
            car_data.append(car_row)

        # Create DataFrame
        calendar_df = pd.DataFrame(car_data, columns=['Car'] + [str(day) for day in range(1, days_in_month + 1)])

        # Convert DataFrame to an OpenPyXL worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(calendar_df, index=False, header=True):
            ws.append(r)

        # Add comments to cells containing booking IDs and change cell color if booked
        for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column)):
            for col_index, cell in enumerate(row):
                if cell.value:  # Check if the cell contains a booking ID
                    if cell.value == 'M':
                        cell.fill = PatternFill(start_color="80FF00", end_color="80FF00", fill_type="solid")  # rgba(128, 255, 0, 0.5)
                    else:
                        booking_id = int(cell.value)
                        booking = Booking.query.get(booking_id)  # Retrieve booking object
                        if booking:
                            # Add comment with booking info
                            booking_info = booking.get_booking_info()
                            comment = Comment(booking_info, 'Booking Details')
                            cell.comment = comment
                            # Change cell color if booked to a lighter shade of green
                            cell.fill = PatternFill(start_color="80FF00", end_color="80FF00", fill_type="solid")  # rgba(128, 255, 0, 0.5)
                

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return the Excel file as a response
        return send_file(
            output,
            download_name=f'calendar_{request.form.get("current_month")}.xlsx',
            as_attachment=True
        )
    
    
    return render_template('calendar.html' if current_user.language == 'en' else f'calendar_{current_user.language}.html', is_current_month=is_current_month,
                            cars=user_cars, booking_data=booking_data, current_month=current_month, current_day=current_day, current_date=current_date, datetime=datetime,
                            days_in_month=days_in_month, user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/contacts', methods=['GET', 'POST'])
@requires_verification
def contacts(): 
    user_contacts = current_user.contacts.all()
    add_and_book_contact = request.args.get('add_and_book_contact')

    if request.method == 'POST':
        if 'full_name' in request.form:
            # Process the form data for adding a contact
            full_name = request.form.get('full_name')
            dob = request.form.get('dob')
            driver_licence_n = request.form.get('driver_licence_n')
            telephone = request.form.get('telephone')
            town_of_birth = request.form.get('town_of_birth')
            city_of_residence = request.form.get('city_of_residence')
            address = request.form.get('address')

            new_contact = Contacts.add_contact(full_name, dob, driver_licence_n, telephone, town_of_birth, city_of_residence, address, current_user.id)
            flash('Contact added successfully', 'success')
            
            if add_and_book_contact:
                return redirect(url_for('overview', contact_id=new_contact.id))                

        elif 'search_type' in request.form:
            # Process the form data for searching a contact
            search_type = request.form.get('search_type')
            search_query = request.form.get('search_query')
            search_results = Contacts.search_contacts(search_type, search_query, current_user.id)
            return render_template('contacts.html' if current_user.language == 'en' else f'contacts_{current_user.language}.html', 
                                   user_contacts=user_contacts, search_query=search_query, search_results=search_results, 
                                   user_name=current_user.username if current_user.is_authenticated else None)
        
        elif 'manage_contact' in request.form:
            # Process the form data for managing a contact
            contact_id = request.form.get('manage_contact')
            return redirect(url_for('contact_manage', contact_id=contact_id))

        elif 'book_contact' in request.form:
            # Process the form data for booking a contact
            contact_id = request.form.get('book_contact')
            return redirect(url_for('overview', contact_id=contact_id))
        
        elif 'delete_contact' in request.form:
            # Handle deletion logic
            contact_id = request.form.get('delete_contact')
            contact = Contacts.query.get_or_404(contact_id)
            db.session.delete(contact)
            db.session.commit()
            flash('Contact deleted successfully', 'success')
            return redirect(url_for('contacts'))

        return redirect(url_for('contacts'))

    return render_template('contacts.html' if current_user.language == 'en' else f'contacts_{current_user.language}.html', 
                           user_contacts=user_contacts, add_and_book_contact=add_and_book_contact if add_and_book_contact else None,
                           user_name=current_user.username if current_user.is_authenticated else None)

    
@app.route('/contact/<int:contact_id>', methods=['GET', 'POST'])
@requires_verification
def contact_manage(contact_id):
    contact = Contacts.query.get_or_404(contact_id)
    contact_bookings = contact.bookings.all()    

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'amend':
            # Handle amendment logic
            contact.full_name = request.form.get('full_name')
            contact.dob = request.form.get('dob')
            contact.driver_licence_n = request.form.get('driver_licence_n')
            contact.telephone = request.form.get('telephone')
            contact.town_of_birth = request.form.get('town_of_birth')
            contact.city_of_residence = request.form.get('city_of_residence')
            contact.address = request.form.get('address')
            db.session.commit()
            flash('Contact details amended successfully', 'success')

        elif action == 'delete':
            # Handle deletion logic
            db.session.delete(contact)
            db.session.commit()
            flash('Contact deleted successfully', 'success')
            return redirect(url_for('contacts'))

    return render_template('contact_manage.html' if current_user.language == 'en' else f'contact_manage_{current_user.language}.html', 
                           contact=contact, contact_bookings=contact_bookings, days=contact.rented_days, 
                           money=contact.money_spent, bookings_number=len(contact_bookings), user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/groups', methods=['GET', 'POST'])
@requires_verification
def groups(): 
    user_groups = current_user.groups.all()

    if request.method == 'POST':
        if 'name' in request.form:
            # Process the form data for adding a group
            name = request.form.get('name')
            telephone = request.form.get('telephone')
            print(name, telephone)
            Groups.add_group(name, telephone, current_user.id)
            flash('Group added successfully', 'success')

        elif 'search_type' in request.form:
            # Process the form data for searching a group
            search_type = request.form.get('search_type')
            search_query = request.form.get('search_query')
            search_results = Groups.search_groups(search_type, search_query, current_user.id)
            return render_template('groups.html' if current_user.language == 'en' else f'groups_{current_user.language}.html', 
                                   user_groups=user_groups, search_query=search_query, search_results=search_results, 
                                   user_name=current_user.username if current_user.is_authenticated else None)
        
        elif 'manage_group' in request.form:
            # Process the form data for managing a group
            group_id = request.form.get('manage_group')
            return redirect(url_for('group_manage', group_id=group_id))
        
        elif 'delete_group' in request.form:
            # Handle deletion logic
            group_id = request.form.get('delete_group')
            group = Groups.query.get_or_404(group_id)
            db.session.delete(group)
            db.session.commit()
            flash('Group deleted successfully', 'success')
            return redirect(url_for('groups'))

        return redirect(url_for('groups'))

    return render_template('groups.html' if current_user.language == 'en' else f'groups_{current_user.language}.html', 
                           user_groups=user_groups, user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/groups/<int:group_id>', methods=['GET', 'POST'])
@requires_verification
def group_manage(group_id):
    group = Groups.query.get_or_404(group_id)
    group_bookings = group.bookings.all()    

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'amend':
            # Handle amendment logic
            group.name = request.form.get('name')
            group.telephone = request.form.get('telephone')
            db.session.commit()
            flash('Group details amended successfully', 'success')

        elif action == 'delete':
            # Handle deletion logic
            db.session.delete(group)
            db.session.commit()
            flash('Group deleted successfully', 'success')
            return redirect(url_for('groups'))

    return render_template('group_manage.html' if current_user.language == 'en' else f'group_manage_{current_user.language}.html', 
                           group=group, group_bookings=group_bookings, money=group.money, bookings_number=group.bookings_number, 
                           user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/downloads', methods=['GET', 'POST'])
@requires_verification
def downloads():
    user_cars = current_user.garage.all()
    user_groups = current_user.groups.all()
    user_contacts = current_user.contacts.all()
    current_date = datetime.now().strftime('%d-%M-%y')    
    alldata_button = request.form.get('allData_button')    
    single_data_button = request.form.get('single_data_button')  

    # Download all time data for: 
    if alldata_button == 'cars':
        if current_user.language == 'it':
            file_name = f'parco_auto_{current_date}.xlsx'
            cars_data = [{'Targa': car.plate, 'Marchio': car.make, 'Modello': car.model, f'{current_user.measurement_unit}': car.km,
                        'Carburante': car.fuel, 'Anno': car.year, 'Cilindrata': car.cc, 'Giorni noleggiati': car.days, 'Ricavo': car.money, 'Costi': car.car_cost, 
                        'Profitto': car.money-car.car_cost, 'Scadenza Assicurazione': car.insurance_expiry_date.strftime('%d-%m-%Y') if car.insurance_expiry_date else None, 
                        'Scadenza Revisione': car.mot_expiry_date.strftime('%d-%m-%Y') if car.mot_expiry_date else None,
                        'Scadenza Bollo': car.road_tax_expiry_date.strftime('%d-%m-%Y') if car.road_tax_expiry_date else None}
                        for car in user_cars]
        else:
            file_name = f'garage_data_{current_date}.xlsx'
            cars_data = [{'Plate': car.plate, 'Make': car.make, 'Model': car.model, f'{current_user.measurement_unit}': car.km,
                        'Fuel': car.fuel, 'Year': car.year, 'CC': car.cc, 'Days rented': car.days, 'Revenue': car.money, 'Costs': car.car_cost, 
                        'Profit': car.money-car.car_cost, 'Insurance expiry': car.insurance_expiry_date.strftime('%d-%m-%Y') if car.insurance_expiry_date else None, 
                        'MOT expiry': car.mot_expiry_date.strftime('%d-%m-%Y') if car.mot_expiry_date else None,
                        'Road Tax expiry': car.road_tax_expiry_date.strftime('%d-%m-%Y') if car.road_tax_expiry_date else None}
                        for car in user_cars]
            
        cars_df = pd.DataFrame(cars_data)
        output = io.BytesIO()
        cars_df.to_excel(output, index=False)
        output.seek(0)
        return send_file(
            output,
            download_name=file_name,
            as_attachment=True
        )
    elif alldata_button == 'bookings':
        user_bookings = Booking.query.filter_by(user_id=current_user.id).all()

        if current_user.language == 'it':
            file_name = f'resoconto_prenotazioni_{current_date}.xlsx'
            bookings_data = [{'Prenotazione ID': booking.id, 'Targa Auto': booking.car.plate, 'Modello Auto': booking.car.model, f'{current_user.measurement_unit}': booking.km,
                            'Prezzo': booking.money, 'Data Inizio': booking.start_datetime.strftime('%d-%m %H:%M'), 'Data Fine': booking.end_datetime.strftime('%d-%m %H:%M'),
                            'Cliente': (Contacts.query.filter_by(id=booking.contact_id).first()).full_name if booking.contact_id else '', 
                            'Gruppo': booking.group.name if booking.group else '', 
                            'Nota': booking.note} for booking in user_bookings]
        else:
            file_name = f'bookings_data_{current_date}.xlsx'
            bookings_data = [{'Booking ID': booking.id, 'Car Plate': booking.car.plate, 'Car Model': booking.car.model, f'{current_user.measurement_unit}': booking.km,
                        'Price': booking.money, 'Start Date': booking.start_datetime.strftime('%d-%m %H:%M'), 'End Date': booking.end_datetime.strftime('%d-%m %H:%M'),
                        'Client': (Contacts.query.filter_by(id=booking.contact_id).first()).full_name if booking.contact_id else '', 
                        'Group': booking.group.name if booking.group else '',
                        'Note': booking.note} for booking in user_bookings]
        bookings_df = pd.DataFrame(bookings_data)
        output = io.BytesIO()
        bookings_df.to_excel(output, index=False)
        output.seek(0)
        return send_file(
            output,
            download_name=file_name,
            as_attachment=True
        )
    elif alldata_button == 'contacts':
        if current_user.language == 'it':
            file_name = f'resoconto_contatti_{current_date}.xlsx'
            contacts_data = [{'ID': contact.id, 'Nome Completo': contact.full_name, 'Numero Patente': contact.driver_licence_n,
                            'Data Di Nascita': contact.dob, 'Telefono': contact.telephone, 'Soldi Spesi': contact.money_spent,
                            'Giorni Noleggiati': contact.rented_days, 'Numero Noleggi': len(contact.bookings.all())  } for contact in user_contacts]
        else:
            file_name = f'contacts_data_{current_date}.xlsx'
            contacts_data = [{'ID': contact.id, 'Full Name': contact.full_name, 'Driver Licence Number': contact.driver_licence_n,
                            'Date Of Birth': contact.dob, 'Telephone': contact.telephone, 'Money Spent': contact.money_spent,
                            'Rented Days': contact.rented_days, 'Bookings Number': len(contact.bookings.all())  } for contact in user_contacts]
        contacts_df = pd.DataFrame(contacts_data)
        output = io.BytesIO()
        contacts_df.to_excel(output, index=False)
        output.seek(0)
        return send_file(
            output,
            download_name=file_name,
            as_attachment=True
        )
    elif alldata_button == 'groups':
        if current_user.language == 'it':
            file_name = f'resoconto_gruppi_{current_date}.xlsx'
            groups_data = [{'ID': group.id, 'Nome Gruppo': group.name, 'Telefono': group.telephone, 'Ricavo': group.money,
                            'Numero Noleggi': group.bookings_number  } for group in user_groups]
        else:
            file_name = f'groups_data_{current_date}.xlsx'
            groups_data = [{'ID': group.id, 'Group Name': group.name, 'Telephone': group.telephone, 'Revenue': group.money,
                            'Bookings Number': group.bookings_number  } for group in user_groups]
        groups_df = pd.DataFrame(groups_data)
        output = io.BytesIO()
        groups_df.to_excel(output, index=False)
        output.seek(0)
        return send_file(
            output,
            download_name=file_name,
            as_attachment=True
        )        
    # Download all time data for: 
    if single_data_button == 'car':
        pass

    return render_template('downloads.html' if current_user.language == 'en' else f'downloads_{current_user.language}.html', 
                           page="downloads", single_data_button=single_data_button if single_data_button else None, user_cars=user_cars,
                           user_name=current_user.username if current_user.is_authenticated else None)

@app.route('/graphs', methods=['GET', 'POST'])
@requires_verification
def graphs():
    '''msg = Message('Oggetto',
                  sender='rentami_team@outlook.com',
                  recipients=['giacomofculcasi@gmail.com'])
    msg.body = 'evviva lo sticchio!'
    mail.send(msg)'''
    user_cars = current_user.garage.all()
    user_contacts = current_user.contacts.all()
    user_groups = current_user.groups.all()

    # To check time frame
    from_date = request.form.get('from')
    to_date = request.form.get('to')

    # Standard time frame if no Input of 1 month
    if from_date == None:
        from_date = datetime.now().strftime('%Y-%m-%d')
        to_date = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
        from_datetime = datetime.strptime(f'{from_date}', '%Y-%m-%d')
        to_datetime = datetime.strptime(f'{to_date}', '%Y-%m-%d')
    # Chosen time frame if Input    
    else:
        from_datetime = datetime.strptime(f'{from_date}', '%Y-%m-%d')
        to_datetime = datetime.strptime(f'{to_date}', '%Y-%m-%d')
        #Include the full to_datetime
        to_datetime += timedelta(days=1)    

    # Fetch all bookings for the user within the specified time range
    user_bookings = Booking.query.filter(
                Booking.user_id == current_user.id,
                Booking.start_datetime >= from_datetime,
                Booking.start_datetime <= to_datetime,
                ).all()

    # Extract the dates of bookings
    booking_dates = [booking.start_datetime.date() for booking in user_bookings]

    # Define the time frame
    time_frame = to_datetime - from_datetime

    # Create a list of all the days in the time frame
    all_dates = [from_datetime + timedelta(days=i) for i in range(time_frame.days + 1)]
    all_dates = [date.date() for date in all_dates]

    # Count the number of bookings for each day
    date_counts = {}
    for date in all_dates:
        date_counts[date] = sum(1 for booking in user_bookings if booking.start_datetime.date() == date)

    # Extract dates and corresponding counts
    dates = list(date_counts.keys())
    counts = list(date_counts.values())

    # Convert the plot data to datetime.date objects
    dates = [date.strftime('%d') for date in dates]

    # Plot the data
    plt.plot(dates, counts, color='blue', marker='', linestyle='-')

    # Customize the plot
    plt.title('Number of Bookings per Day')
    plt.xlabel('Day')
    plt.ylabel('Number of Bookings')
    plt.xticks(rotation=45, ha='right')

    # Show the plot
    plt.tight_layout()

    # Save the plot to a buffer and convert to base64
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plot_data = base64.b64encode(buffer.getvalue()).decode()
    plt.close()  # Close the plot to prevent it from displaying in the console


    return render_template('graphs.html' if current_user.language == 'en' else f'graphs_{current_user.language}.html',
                           from_datetime=from_datetime.strftime('%Y-%m-%d'), to_datetime=to_datetime.strftime('%Y-%m-%d'),
                           user_name=current_user.username if current_user.is_authenticated else None,
                           plot_data=plot_data)